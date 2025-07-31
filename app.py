from flask import Flask, request, send_file,render_template
import pandas as pd
import tempfile, os, datetime, re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


app = Flask(__name__)
app.secret_key = "replace-me"

# ──────────────────────────────────────────────────────────────
# 0 ▸ helpers
# ──────────────────────────────────────────────────────────────
def norm_pn(text: str) -> str:
    """Upper-case & keep only A–Z, 0–9."""
    return re.sub(r"[^A-Z0-9]", "", str(text).upper())

def sanitise_quantity(df: pd.DataFrame, qty_col: str) -> pd.DataFrame:
    qty = pd.to_numeric(df[qty_col], errors="coerce").fillna(1)
    df[qty_col] = qty.where(qty > 0, 1).astype(int)
    return df

# ──────────────────────────────────────────────────────────────
# 1 ▸ INVENTORY  (List Coder CSV)
# ──────────────────────────────────────────────────────────────
def load_inventory(fileobj):
    raw = pd.read_csv(fileobj, dtype=str).fillna("")
    qty_col = "inscor__Quantity_Available__c"
    raw = raw[pd.to_numeric(raw[qty_col], errors="coerce").fillna(0).astype(int) > 0]
    raw = sanitise_quantity(raw, qty_col)

    def best_price(r):
        try:
            p = float(r["SSP_Updated__c"])
            return p if p > 0 else float(r["Extended_SSP__c"])
        except Exception:
            return ""
    raw["$PRICE"] = raw.apply(best_price, axis=1)

    keep = {
        "QUANTITY"      : qty_col,
        "PART NUMBER"   : "inscor__Product__r.Name",
        "SERIAL NUMBER" : "inscor__Serial_Number__c",
        "CONDITION"     : "inscor__Condition_Code__r.Name",
        "TAG DATE"      : "inscor__Tag_Date__c",
        "TAGGED BY ONE" : "Tag_Agency_Merged__c",
        "LAST OPERATOR" : "inscor__Trace__r.Name",
        "SALES PRICE"   : "$PRICE",
    }
    inv = pd.DataFrame({k: raw[v] for k, v in keep.items()})
    inv["LEAD TIME"]    = ""
    inv["TAGGED BY"]    = ""
    inv["CERTIFICATES"] = ""
    inv["CURRENCY"]     = "USD"
    return inv

# ──────────────────────────────────────────────────────────────
# 2 ▸ REQUESTS  (Received Requests XLSX)
# ──────────────────────────────────────────────────────────────
def load_requests(fileobj):
    req = pd.read_excel(fileobj, dtype=str).fillna("")

    # — drop stray “Remarks” column if it shows up at position F —
    if len(req.columns) > 5 and req.columns[5].strip().lower() == "remarks":
        req.drop(columns=req.columns[5], inplace=True)
        req.reset_index(drop=True, inplace=True)

    # locate the comma-list “Condition” column (usually col I)
    cond_idx = 8 if len(req.columns) > 8 else len(req.columns) - 1
    cond_col = req.columns[cond_idx]

    split = (req[cond_col]
             .str.split(",", expand=True).iloc[:, :4]
             .fillna("").applymap(str.strip))
    for i in range(split.shape[1]):
        req[f"COND_{i+1}"] = split[i]
    return req

# ──────────────────────────────────────────────────────────────
# 3 ▸ LEAD-TIME flag (strict PN + Condition match)
# ──────────────────────────────────────────────────────────────
def set_lead_time(inv, req):
    inv["PN_N"]   = inv["PART NUMBER"].apply(norm_pn)
    inv["COND_N"] = inv["CONDITION"].str.strip().str.upper()

    pn_col    = req.columns[5]           # Part No. column (still F)
    cond_cols = [c for c in req.columns if c.startswith("COND_")]
    req_long  = (req.melt(id_vars=[pn_col], value_vars=cond_cols, value_name="COND")
                    .dropna(subset=["COND"]).query("COND != ''"))
    req_long["PN_N"]   = req_long[pn_col].apply(norm_pn)
    req_long["COND_N"] = req_long["COND"].str.strip().str.upper()
    valid = set(zip(req_long["PN_N"], req_long["COND_N"]))

    inv["LEAD TIME"] = [
        1 if (pn, cd) in valid else ""
        for pn, cd in zip(inv["PN_N"], inv["COND_N"])
    ]
    return inv.drop(columns=["PN_N", "COND_N"])

# ──────────────────────────────────────────────────────────────
# 4 ▸ CERTIFICATES & TAGGED-BY merge
# ──────────────────────────────────────────────────────────────
def fill_certificates(inv):
    def cert(cond):
        cond = cond.upper()
        if cond in ("SV", "OH"):  return "FAA,EASA"
        if cond in ("NE", "NEW", "FN"): return "FAA"
        return ""
    inv["CERTIFICATES"] = inv["CONDITION"].apply(cert)
    inv.loc[inv["TAGGED BY"] == "", "TAGGED BY"] = inv["TAGGED BY ONE"]
    return inv.drop(columns=["TAGGED BY ONE"])

# ──────────────────────────────────────────────────────────────
# 5 ▸ UNMATCHED REQUEST sheet
# ──────────────────────────────────────────────────────────────
def unmatched_req_sheet(inv_all, req):
    have = set(inv_all["PART NUMBER"].apply(norm_pn))
    want = req[req.columns[5]].apply(norm_pn)
    missing = want[~want.isin(have)].drop_duplicates().rename("Part Number")
    return missing.to_frame()

# ──────────────────────────────────────────────────────────────
# 6 ▸ Attach request meta + placeholders
# ──────────────────────────────────────────────────────────────
def attach_request_meta(df, req):
    lookup = {norm_pn(r.iloc[5]): r for _, r in req.iterrows()}
    for col in ["Request No.","Delivery Date","Demander","Transaction Types"]:
        df.insert(0, col, "")

    for i, r in df.iterrows():
        key = norm_pn(r["PART NUMBER"])
        if key in lookup:
            src = lookup[key]
            df.at[i, "Request No."]       = src.iloc[0]
            df.at[i, "Delivery Date"]     = src.iloc[1]
            df.at[i, "Demander"]          = src.iloc[2]
            df.at[i, "Transaction Types"] = src.iloc[3]

    for col in ["Manufacturer","Last MSN","Deviations","Comment",
                "Expiration Date","Loan Price","Exchange Price"]:
        if col not in df.columns:
            df[col] = ""
    return df

def uplift_price(df):
    df["SALES PRICE"] = pd.to_numeric(df["SALES PRICE"], errors="coerce").fillna(0) * 1.10
    return df

FINAL_ORDER = [
    "Request No.","Delivery Date","Demander","Transaction Types",
    "QUANTITY","PART NUMBER","SERIAL NUMBER","Manufacturer",
    "CONDITION","LEAD TIME","TAGGED BY","TAG DATE","CERTIFICATES",
    "Expiration Date","Comment","Deviations","Last MSN",
    "LAST OPERATOR","SALES PRICE","CURRENCY",
    "Loan Price","Exchange Price"
]

# ──────────────────────────────────────────────────────────────
# 7 ▸ Row colour (TAG DATE age)
# ──────────────────────────────────────────────────────────────
def colour_rows(path):
    wb = load_workbook(path)
    ws = wb["Match value"]
    headers = [c.value for c in ws[1]]
    try:
        tag_idx = headers.index("TAG DATE")
    except ValueError:
        wb.save(path); return
    fills = {
        "red"   : PatternFill("solid","FF0000"),
        "green" : PatternFill("solid","00FF00"),
        "yellow": PatternFill("solid","FFFF00"),
    }
    today = datetime.date.today()
    for row in ws.iter_rows(min_row=2, values_only=False):
        cell = row[tag_idx]
        if not cell.value: continue
        try:
            d = (cell.value if isinstance(cell.value, datetime.date)
                 else datetime.datetime.strptime(str(cell.value)[:10], "%Y-%m-%d").date())
            diff = (today.year - d.year)*12 + today.month - d.month
            fill = fills["red"] if diff > 24 else fills["green"] if diff >= 8 else fills["yellow"]
            for c in row: c.fill = fill
        except Exception:
            pass
    wb.save(path)

# ──────────────────────────────────────────────────────────────
# 8 ▸ Flask routes
# ──────────────────────────────────────────────────────────────
HTML_FORM = """<!doctype html><title>Lufty Coder</title>
<h2>Lufty Coder Converter</h2>
<form action="/process" method="post" enctype="multipart/form-data">
  List Coder CSV : <input type="file" name="lc" accept=".csv"><br><br>
  Requests XLSX  : <input type="file" name="rq" accept=".xlsx"><br><br>
  <button>Convert</button>
</form>"""

@app.route("/", methods=["GET"])
def index():
    return render_template("index.html", year=datetime.date.today().year)

@app.route("/process", methods=["POST"])
def process():
    lc = request.files.get("lc"); rq = request.files.get("rq")
    if not lc or not rq:
        return "Both files required", 400
    try:
        inv = load_inventory(lc)
        req = load_requests(rq)
        inv = set_lead_time(inv, req)
        inv = fill_certificates(inv)

        match_df = inv[inv["LEAD TIME"] == 1].copy()
        match_df = attach_request_meta(match_df, req)
        match_df = uplift_price(match_df)
        match_df = match_df[FINAL_ORDER]        # exact order

        unmatched_df = unmatched_req_sheet(inv, req)

        fd, path = tempfile.mkstemp(suffix=".xlsx"); os.close(fd)
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            match_df.to_excel(writer,    sheet_name="Match value",          index=False)
            unmatched_df.to_excel(writer, sheet_name="Not Matching Values", index=False)
        colour_rows(path)
        return send_file(path, as_attachment=True, download_name=rq.filename)
    except Exception as e:
        return f"Error: {e}", 500

if __name__ == "__main__":
    app.run(debug=True)
